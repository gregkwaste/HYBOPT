import os
import openpyxl
import re
import sys


# Define the paths
#solution_folder = r'D:\Repos\MSOP\MSOP\extracted_solutions\Random51'
#solution_folder = '..\extracted_solutions\selected51_CPX'
#solution_folder = '..\extracted_solutions\selected51_MLS'
#solution_folder = '..\extracted_solutions\selected51_GRB'

# MSOP unit clusters
#solution_folder = '..\extracted_solutions\Experiment mls_unit_clusters\\unit'
solution_folder = '..\extracted_solutions\Experiment mls_unit_clusters\\normal'

method = 'mls' # exact / mls

excel_file_path = os.path.join(solution_folder, 'results.xlsx')

# Open the Excel file
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook['results']

# Write the headers
if method == 'exact':
    headers = ['Num', 'Instance', 'Data', 'T', 'P', 'V', 'Sets', 'Nodes', 'Profit', 'Time', 'Status', 'NumNodes', 'BB', 'BBRoot', 'BBRootBefCuts', 'Root Gap', 'Final Gap']
elif method == 'mls':
    headers = ['Num', 'Instance', 'Data', 'T', 'P', 'V', 'Sets', 'Nodes', 'Profit', 'Time', 'Status', 'BestBound', 'BestRestart', 'Restart time', 'Total iterations', 'Iterations till Best', 'Math time', 'Final Gap']

for col, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col).value = header

# Initialize the row counter
row = 2

# Loop through the solution files
for filename in os.listdir(solution_folder):
    if filename.endswith('.txt'):
        # Open the solution file
        with open(os.path.join(solution_folder, filename), 'r') as file:
            # Read the data
            dataset_name = file.readline().split(': ')[1].strip()
            vehicles = file.readline().split(': ')[1].strip()
            routes = []
            line = file.readline().strip() # Routes
            while True:
                line = file.readline().strip()
                if line.startswith('Route_'):
                    routes.append(line.split(': ')[1].split(','))
                else:
                    profit = float(line.split(': ')[1].strip())
                    break
            #profit = file.readline().split(': ')[1].strip()
            time = float(file.readline().split(': ')[1].strip())
            file.readline()

            status = file.readline().split(': ')[1].strip()
            if method == 'exact':
                num_nodes = int(file.readline().split(': ')[1].strip())
                best_bound = float(file.readline().split(': ')[1].strip())
                best_bound_root = float(file.readline().split(': ')[1].strip())
                best_bound_root_before_cuts = float(file.readline().split(': ')[1].strip())
                final_gap = float(file.readline().split(': ')[1].strip())
            elif method == 'mls':
                best_bound = float(file.readline().split(': ')[1].strip())
                best_restart = float(file.readline().split(': ')[1].strip())
                restart_time = float(file.readline().split(': ')[1].strip())
                total_iters = float(file.readline().split(': ')[1].strip())
                iter_till_best = float(file.readline().split(': ')[1].strip())
                time_of_math = float(file.readline().split(': ')[1].strip())

        # Write the data to the Excel file
        sheet.cell(row=row, column=1).value = row - 1
        sheet.cell(row=row, column=2).value = dataset_name

        parts = dataset_name.split('_')

        sheet.cell(row=row, column=3).value = parts[0]
        sheet.cell(row=row, column=4).value = parts[1]
        sheet.cell(row=row, column=5).value = parts[2]
        sheet.cell(row=row, column=6).value = parts[3]

        #match = re.match(r'^(\d+)(.*)?(\d+)?$', parts[0])
        ##sets = int(match.group(1))
        #customers = int(match.group(3))
        sets, customers = re.findall(r'\d+', parts[0])
        sheet.cell(row=row, column=7).value = sets
        sheet.cell(row=row, column=8).value = customers
        sheet.cell(row=row, column=9).value = profit
        sheet.cell(row=row, column=10).value = time
        sheet.cell(row=row, column=11).value = status

        if method == 'exact':
            sheet.cell(row=row, column=12).value = num_nodes
            sheet.cell(row=row, column=13).value = best_bound
            sheet.cell(row=row, column=14).value = best_bound_root
            sheet.cell(row=row, column=15).value = best_bound_root_before_cuts
            rootgap = ''
            final_gap = ''
            if profit != 0:
                rootgap = 100 * ((best_bound_root-profit)/profit)
                final_gap = 100 * ((best_bound-profit)/profit)
            sheet.cell(row=row, column=16).value = rootgap
            sheet.cell(row=row, column=17).value = final_gap
        elif method == 'mls':
            sheet.cell(row=row, column=12).value = best_bound
            sheet.cell(row=row, column=13).value = best_restart
            sheet.cell(row=row, column=14).value = restart_time

            sheet.cell(row=row, column=15).value = total_iters
            sheet.cell(row=row, column=16).value = iter_till_best
            sheet.cell(row=row, column=17).value = time_of_math
            final_gap = ''
            if profit != 0:
                final_gap = 100 * ((best_bound-profit)/best_bound)
            sheet.cell(row=row, column=18).value = final_gap

        # Increment the row counter
        row += 1

# Save the Excel file
workbook.save(excel_file_path)